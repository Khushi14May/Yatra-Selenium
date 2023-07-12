import openpyxl


class HomePageData:

    # @staticmethod
    # def getTestData(test_case_name):
    #     Dict = {}
    #
    #     book = openpyxl.load_workbook("C:\\Users\\158421\\OneDrive - Arrow Electronics, Inc\\Desktop\\Yatra.xlsx")
    #     sheet = book.active
    #     for i in range(1, sheet.max_row + 1):  # to get rows
    #         if sheet.cell(row=i, column=1).value == test_case_name:
    #
    #             for j in range(2, sheet.max_column + 1):  # to get columns
    #                 Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
    #     return[Dict]

#used for having more then one senarios of particular testcase
    @staticmethod
    def getTestData(test_case_name):
        global l1
        list_data = []
        Dict = {}
        l1=[]
        book = openpyxl.load_workbook("C:\\Users\\158421\\OneDrive - Arrow Electronics, Inc\\Documents\\UdemyAppium\\Yatra\\Utilities\\Yatra.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=2).value == test_case_name:
                l1.append(i)
        # print(l1)
        for k in l1:
            for j in range(3, sheet.max_column + 1):  # to get columns
                Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=k, column=j).value
            list_data.append(Dict)
            Dict={}
        # print(list_data)
        return list_data


