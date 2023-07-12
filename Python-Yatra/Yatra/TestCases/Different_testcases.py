import openpyxl
from TestCases.test_Master import Test_Page

class test_scripts:
    def getTestData(self):
        global list_data
        Dict = {}
        l1 = []
        list_data = []
        book = openpyxl.load_workbook("C:\\Users\\158421\\OneDrive - Arrow Electronics, Inc\\Documents\\UdemyAppium\\Yatra\\Utilities\\Yatra.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == 'y':
                l1.append(i)
        for k in l1:
            for j in range(2, sheet.max_column + 1):  # to get columns
                Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=k, column=j).value
            list_data.append(Dict)
            Dict = {}
        return list_data
        # print(len(list_data))



    def Call_testcase(self,data):
        global str_case1
        for i in range(len(data)):
            dict1=data[i]
            print("dict")
            print(dict1)
            testcase=dict1['Testcase']
            print("test"+testcase)
            str_case1="test_"+testcase
            hp=Test_Page()
            getattr(hp, str_case1)(dict1)
            dict1={}


obj=test_scripts()
Dict=obj.getTestData()
obj.Call_testcase(Dict)